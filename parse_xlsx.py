#XLSX Parser - it parses an XLSX file for cell values
import openpyxl

workb = openpyxl.load_workbook('layer.xlsx')  # run from CWD where XLSX lives
sheet = workb[workb.sheetnames[0]] # the sheet to parse

TTP = []  # Assigning the cell values to this list

def loop_thru(sheet):  # loops through all cells in worksheet
    from openpyxl import worksheet
    R = 2 # starting with row 2 bc first row is tactic header
    C = 1
    while C <= int(sheet.max_column):
        for i in range(1, int(sheet.max_row)): 
            val = sheet.cell(row=R, column=C).value
            if val == None:  # None means this is the end of that column
                C = C+1
                R = 2
            else:
                TTP.append(val)
                R = R+1

def write_text(TTP: list):
    with open('TTP.txt', 'a') as writer:
        for i in TTP:
            writer.write((i+'\n'))
        writer.close()


def write_list(TTP: list):
    with open('TTP_list.txt', 'a') as writer:
        writer.write(str(TTP))
        writer.close()

loop_thru(sheet)
TTP = list(set(TTP)) # remove duplicates
TTP.sort() # alphabetical
write_text(TTP) # write to .txt
write_list(TTP) # write the list to another file for later use
print(TTP)  # prints list to stdout
